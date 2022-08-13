import pandas as pd
import requests
import io
import streamlit as st

from diskcache import Cache
from openpyxl import load_workbook


def load_data():
    df = pd.read_csv("./microdados_censo_escolar_2021/2021/dados/microdados_ed_basica_2021.csv",  sep=';', encoding="latin-1")
    df['CO_ORGAO_REGIONAL'] = df['CO_ORGAO_REGIONAL'].astype(str)
    df['NU_ENDERECO'] = df['NU_ENDERECO'].astype(str)
    df['NU_DDD'] = df['NU_DDD'].astype('Int64').astype('str')
    df['NU_TELEFONE'] = df['NU_TELEFONE'].astype('Int64').astype('str')
    return df

@st.cache
def load_data_pd():
    with open("./microdados_censo_escolar_2021/2021/dados/microdados_ed_basica_2021.txt", "r", encoding="utf-8") as f:
        link = f.readline()
        print("link é " + link)

    df = pd.read_csv(io.BytesIO(requests.get(link).content),  sep=';', encoding="latin-1")
    df['CO_ORGAO_REGIONAL'] = df['CO_ORGAO_REGIONAL'].astype(str)
    df['NU_ENDERECO'] = df['NU_ENDERECO'].astype(str)
    df['NU_DDD'] = df['NU_DDD'].astype('Int64').astype('str')
    df['NU_TELEFONE'] = df['NU_TELEFONE'].astype('Int64').astype('str')
    
    geocache = Cache("geo.cache")

    df["lat"] = 0
    df["lng"] = 0

    for cep in geocache:
        location = geocache[cep][0]["geometry"]["location"]
        df.loc[df["CO_CEP"] == cep, "lat"] = location["lat"]
        df.loc[df["CO_CEP"] == cep, "lng"] = location["lng"]

    return df

@st.cache
def load_helpers():
    wb = load_workbook(filename="./microdados_censo_escolar_2021/2021/Anexos/ANEXO I - Dicionário de Dados/dicionário_dados_educação_básica.xlsx")
    ws = wb["Cadastro_Escolas"]
    
    helpers = {}
    for cellB, cellC, cellF in zip(ws["B"], ws["C"], ws["F"]):
        helpers[cellB.value] = cellC.value
        if cellF.value is not None:
            helpers[cellB.value] += "\n" + cellF.value
    
    for column in ["Alteração de nomenclatura", "Variável nova", "Descontinuidade", None, "Nome da Variável"]:
        del helpers[column]

    return helpers
    
    